import openpyxl, io, json, os, re
root = r"D:/download/pic-vec/feihe/feihe-mvp"
src = os.path.join(root, "data", "qicui", "supplier-comments.xlsx")
wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
def norm_date(name):
    m = re.search(r"(\d{1,2})\.(\d{1,2})", name)
    if not m:
        return None
    return "%02d-%02d" % (int(m.group(1)), int(m.group(2)))
def kind_of(name):
    if "素人" in name:
        return "suren"
    if "达人" in name:
        return "daren"
    return "other"
dates = {}
for name in wb.sheetnames:
    d = norm_date(name)
    if not d or name == "total":
        continue
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        continue
    hdr = [(str(c) if c is not None else "") for c in rows[0]]
    def col(*keys):
        for i, h in enumerate(hdr):
            for k in keys:
                if k in h:
                    return i
        return -1
    c_link = col("笔记链接", "链接")
    c_text = col("评论话术", "话术")
    c_form = col("评论形式")
    c_reply = col("盖楼回复", "楼中楼形式")
    c_nick = col("博主昵称", "达人昵称")
    c_brand = col("品牌侧")
    c_prod = col("产品侧")
    if c_link < 0 or c_text < 0:
        continue
    notes = {}
    for r in rows[1:]:
        link = str(r[c_link]).strip() if r[c_link] else ""
        txt = str(r[c_text]).strip() if r[c_text] else ""
        if not link or link == "None" or not txt or txt == "None":
            continue
        g = notes.get(link)
        if g is None:
            nick = str(r[c_nick])[:24] if c_nick >= 0 and r[c_nick] else ""
            g = {"link": link, "blogger": nick, "sheets": [], "count": 0, "forms": {}, "samples": []}
            notes[link] = g
        g["count"] += 1
        if name not in g["sheets"]:
            g["sheets"].append(name)
        f = str(r[c_form])[:16] if c_form >= 0 and r[c_form] else ""
        g["forms"][f] = g["forms"].get(f, 0) + 1
        if len(g["samples"]) < 6:
            g["samples"].append({"t": txt[:120], "f": f, "r": (str(r[c_reply])[:16] if c_reply >= 0 and r[c_reply] else ""), "b": (str(r[c_brand])[:20] if c_brand >= 0 and r[c_brand] else ""), "p": (str(r[c_prod])[:20] if c_prod >= 0 and r[c_prod] else "")})
    e = dates.get(d)
    if e is None:
        e = {"sheets": [], "notes": []}
        dates[d] = e
    e["sheets"].append({"name": name, "kind": kind_of(name), "rows": len(rows) - 1, "notes": len(notes)})
    e["notes"].extend(list(notes.values()))
for d, e in dates.items():
    merged = {}
    for n in e["notes"]:
        m = merged.get(n["link"])
        if m is None:
            merged[n["link"]] = n
        else:
            m["count"] += n["count"]
            m["sheets"].extend([s for s in n["sheets"] if s not in m["sheets"]])
            m["samples"].extend(n["samples"][:2])
            m["samples"] = m["samples"][:6]
    e["notes"] = sorted(merged.values(), key=lambda x: -x["count"])
    e["noteCount"] = len(e["notes"])
    e["rowCount"] = sum(n["count"] for n in e["notes"])
out = {"source": "supplier-comments.xlsx", "dates": dates}
p = os.path.join(root, "data", "qicui", "supplier-seed.json")
io.open(p, "w", encoding="utf-8").write(json.dumps(out, ensure_ascii=False))
print("dates", len(dates), "bytes", os.path.getsize(p))
for d in sorted(dates)[:5]:
    print(d, dates[d]["noteCount"], dates[d]["rowCount"])
